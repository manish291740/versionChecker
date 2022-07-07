from django.shortcuts import render,redirect
from .forms import checkDataModel,checkForm
import re
# sender_email = "yogesh.gaur@thysserkrupp.com"
# password = "@123YogiMeraNaam"

# Create your views here.
# import smtplib, ssl,re
#
# def mail(ip):
#     smtp_server = "smtp.gmail.com"
#     port = 465  # For start tls
#     sender_email = "pubgkoreandj@gmail.com"
#     password = "rsyxnvyhwgrzhxtj"
#     receiver_email = "dheerajap.a99@outlook.com"
#     # Create a secure SSL context
#
#     SUBJECT = "System Readiness Check IP Address"
#     TEXT = f"""Hello Team,
#
#                 This Machine name is send for Processed
#                 --{ip}--
#
#
#                 Regards,
#                 Bot"""
#
#     message = 'Subject: {}\n\n{}'.format(SUBJECT, TEXT)
#     context = ssl.create_default_context()
#     with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
#         server.login(sender_email, password)
#         server.sendmail(sender_email, receiver_email, message)

def Checker(request):
    if request.method=="POST":
        ShowData=request.POST.get("Ip_Address")
        Show = ShowData.split(',')
        errormsg = " is not an Machine Name."
        store=[]
        errorList=[]
        for i in set(Show):
            if re.fullmatch(r'^[A-Za-z][\S]{3}[0-9][\S]{4}(.in.uhde.org)', i):
                store.append(i)
            else:
                errorList.append(i)


            # try: Tkmu80097.in.uhde.org, Tkmu61469.in.uhde.org
            #     if type(eval(str(i).replace('.',''))) == int and '.' in str(ShowData):
            #             # mail(i)
            #             store.append(i)#list type
            #             # return render(request,'Show.html',{'show':store})
            #     else:
            #         errorList.append(i)
            # except:
            #     errorList.append(i)
                # return render(request, 'error.html', {'err': str(Show) + errorList})
        # storestr=str(store)
        # import xlsxwriter

        # workbook = xlsxwriter.Workbook('C:\\Automation Anywhere\\Thyssen\\Bot_File_Trigger\\IpAddressData.xlsx')
        # worksheet = workbook.add_worksheet()

        # bold = workbook.add_format({'bold': True})

        # worksheet.write('A1', 'SR.NO', bold)
        # worksheet.write('B1', 'IP Address', bold)
        # row = 1
        # col = 0
        # for ip in store:
        #     print(ip)

        import openpyxl
        from openpyxl.styles import Font
        import os.path
        import datetime
        current_time = datetime.date.today()
        todayDate = str(current_time)
        new_currentDate = todayDate.replace(":", "-")
        filePath_machine_name = (f"C:\Automation Anywhere\Thyssen\Bot_File_Trigger\IpAddressData--{new_currentDate}.xlsx")


        if os.path.isfile(filePath_machine_name):
            workbook_obj = openpyxl.load_workbook(filePath_machine_name)
            sheet_obj = workbook_obj.active
            max_row_count = sheet_obj.max_row + 1
            for ipdata in store:
                sheet_obj.cell(row=max_row_count, column=1).value = ipdata
                sheet_obj.cell(row=max_row_count, column=2).value = "Pending"
                max_row_count = max_row_count + 1
            workbook_obj.save(filePath_machine_name)
        else:

            New_wb_obj=openpyxl.Workbook()
            New_wb_obj.save(filePath_machine_name)
            workbook_obj = openpyxl.load_workbook(filePath_machine_name)
            sheet_obj = workbook_obj.active
            sheet_obj['A1'] = "Machine Name"
            sheet_obj['B1'] = "Status"
            sheet_obj['A1'].font = Font(bold=True)
            sheet_obj['B1'].font = Font(bold=True)

            max_row_count = sheet_obj.max_row + 1
            for ipdata in store:
                sheet_obj.cell(row=max_row_count, column=1).value = ipdata
                sheet_obj.cell(row=max_row_count, column=2).value = "Pending"
                max_row_count = max_row_count + 1
            workbook_obj.save(filePath_machine_name)

        # mail(storestr.replace("'","").replace('[','').replace(']',''))
        # mail(store)
        return render(request,'Show.html',{'show':store,'err': str(errorList) + errormsg,'errorLi':errorList})
    else:
        return render(request, 'HomeIP.html')

